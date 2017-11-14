#!/usr/bin/env python3

import openpyxl

import os.path
import re
import sys
import uuid


def check_value(value, dtype):
    try:
        dtype(value)
        return value
    except Exception:
        return ""


def convert_value(value, dtype):
    try:
        return dtype(value)
    except Exception:
        return dtype()


def load_workbooks(dirs):
    for folder in dirs:
        for filename in os.listdir(folder):
            fn = os.path.join(folder, filename)
            yield ((folder, filename), openpyxl.load_workbook(fn))


def create_export_skeleton():
    SCHEMA = (
        ('ООП', [
            'ИД_ООП', 'УГН', 'Направление', 'Ступень ВО', 'Форма']),
        ('Участник обучения', [
            'ИД_Участ', 'ИД_ООП', 'Основание поступления',
            'Форма финансирования', 'Форм. подразделение',
            'Номер курса обучения', 'Год поступления', 'Срок обучения',
            'Средний балл студента', 'Медиана балла потока',
            'Стд. откл. балла потока', 'Мода балла потока']),
        ('Опросник студента подход 1', [
            'ИД_Участ', 'Дата тестирования', 'Пол', 'Возраст',
            'Место проживания', 'Уровень прокрастинации 1',
            'Уровень прокрастинации 2', 'Уровень прокрастинации 3',
            'Уровень прокрастинации 4', 'Уровень прокрастинации 5',
            'Уровень прокрастинации 6', 'Уровень прокрастинации 7',
            'Уровень прокрастинации 8', 'Уровень прокрастинации 9',
            'Уровень прокрастинации 10',
            'Уровень сам. рег. обучения и владения метакогн. способностями 1',
            'Уровень сам. рег. обучения и владения метакогн. способностями 2',
            'Уровень сам. рег. обучения и владения метакогн. способностями 3',
            'Уровень сам. рег. обучения и владения метакогн. способностями 4',
            'Уровень сам. рег. обучения и владения метакогн. способностями 5',
            'Уровень сам. рег. обучения и владения метакогн. способностями 6',
            'Уровень сам. рег. обучения и владения метакогн. способностями 7',
            'Уровень сам. рег. обучения и владения метакогн. способностями 8',
            'Уровень сам. рег. обучения и владения метакогн. способностями 9',
            'Уровень сам. рег. обучения и владения метакогн. способностями 10',
            'Учебная мотивация 1', 'Учебная мотивация 2',
            'Учебная мотивация 3', 'Учебная мотивация 4',
            'Учебная мотивация 5', 'Учебная мотивация 6',
            'Учебная мотивация 7', 'Учебная мотивация 8',
            'Учебная мотивация 9', 'Учебная мотивация 10',
            'Уровень вовлеченности 1', 'Уровень вовлеченности 2',
            'Уровень вовлеченности 3', 'Уровень вовлеченности 4',
            'Уровень вовлеченности 5', 'Уровень вовлеченности 6',
            'Уровень вовлеченности 7', 'Уровень вовлеченности 8',
            'Уровень вовлеченности 9', 'МООС платформа',
            'Наличие опыта освоения elearning', 'Последствия не освоения',
            'Наличие очных встреч']),
        ('Опросник студента подход 2', [
            'ИД_Участ', 'Дата тестирования', 'Университет', 'Уровень обучения',
            'Направление подготовки', 'Курс', 'Успеваемость за последний год',
            'Форма финансирования', 'email', 'Фамилия', 'Имя', 'Отчество',
            'Выбранный курс', 'Причина выбора курса',
            'Уровень самостоятельности', 'Формат обучения',
            'Формат процесса освоения дисц.', 'Формат итоговой аттестации',
            'Виды активности во время курса', 'Сложность курса',
            'Нагрузка во время курса', 'Оценка контенте курса',
            'Соответствие полученной оценки реальны знаниям',
            'Удовлетворенность всем курсом',
            'Удовлетворенность отдельными элементами курса',
            'Полнота представленности контента по курсу', 'Оценка курсов НПОО',
            'Уровень подготовки до начала обучения',
            'Уровень подготовки полсе освоения курса',
            'Количество времени, затрачиваемого на курс',
            'Нацеленность на получение сертификата', 'Перезачет дисциплины',
            'Степень готовности к итоговому тесту',
            'Предпочтительный формат обучения',
            'Предыдущий опыт онлайн-обучения']),
        ('Данные за семестр', [
            'ИД_Дан_Сем', 'ИД_Участ', 'Семестр', 'Средняя оценка',
            'Средняя оценка(по 1ой сдаче)', 'Количество пересдач']),
        ('Деятельность', [
            'ИД_Деят', 'ИД_Участ', 'Тип (ВО, СО, Олимпиада, ЕГЭ, Вступ.)',
            'Дата начала', 'Дата окончания',
            'Название обр.уч./Уровень олимпиады', 'Тип СО', 'Тип Олимпиады',
            'Название предмета', 'Оценка (ЕГЭ) или средний бал(ВО, СО)']),
        ('Опрос преподавателя', [
            'ИД_Преподавателя', 'Пол', 'Возраст',
            'Уровень образования или уч.степень',
            'Совокупный преподавательский стаж', 'Стаж преподавания в ВУЗе',
            'Преподавательская должность', 'Научная должность',
            'Административная должность', 'Преподавтаельская ставка',
            'Опыт повышения квалификации', 'Форма контроля',
            'Вспомогательные материалы', 'Обратная связь']),
        ('Дисциплина из УП', [
            'ИД_Дисц_УП', 'Название модуля', 'Название', 'Часть УП',
            'Число з.е.', 'Число часов']),
        ('Дисциплина с платформы', [
            'ИД_Дисц_платформы', 'Название модуля', 'Количество недель',
            'Число з.е.', 'Кол-во промежуточных тестирований',
            'Наличие эссе заданий', 'Требования начального уровня',
            'Тематическая направленность']),
        ('Участн-Курс-Дисциплина', [
            'ИД_Участн_Дисц_Курс', 'ИД_Участ', 'ИД_Дисц_УП',
            'ИД_Дисц_платформы', 'ИД_Преподавателя', 'Логин',
            'ид_пользоват_в курсе', 'Модель обучения', 'is_online',
            'Поставщик']),
        ('Результаты обучения на курсе', [
            'ИД_Результат', 'ИД_Участн_Дисц_Курс', 'Вес текущей',
            'Вес промежуточной', 'Балл за промежуточную', 'Балл за текущую',
            'Медиана балла потока по дисциплине', 'Мода балла по дисциплине',
            'Стд откл балла по дисциплине']),
        ('Текущая успеваемость', [
            'ИД_Мероприятия', 'ИД_Результат', 'Название мероприятия',
            'Тип(Лекция, Тест)', 'Время начала', 'Время окончания',
            'Балл за мероприятие', 'Вес мероприятия']),
        ('Задание', [
            'ИД_Задания', 'ИД_Мероприятия', 'текст задания', 'вариант',
            'текстовое значение ответа', 'метка времени',
            'дата и время выгрузки', 'максимальный бал за задание',
            'бал за задание']),
    )

    workbook = openpyxl.Workbook()
    workbook.remove_sheet(workbook.worksheets[0])

    for (name, columns) in SCHEMA:
        worksheet = workbook.create_sheet(title=name)
        worksheet.append(columns)

    return workbook


class StudentsFactory:
    def __init__(self):
        self.students = {}

    def get_or_create(self, *key):
        has = key in self.students
        if not has:
            self.students[key] = str(uuid.uuid4())
        return (self.students[key], not has)


class WorksheetLegend:
    def __init__(self, row):
        self.index = {}
        self.names = []

        for (i, cell) in enumerate(row):
            self.index[cell.value] = i
            self.names.append(cell.value)

    def get_item(self, row, *keys, dtype=lambda x: x):
        for key in keys:
            try:
                return dtype(row[self.index[key]].value)
            except KeyError:
                pass

        return ""


def get_tasks_edx_urfu(worksheet):
    skip_head = [
        'id', 'email', 'username', 'full_name', 'second_name', 'grade',
        'Student ID', 'Email', 'Username', 'Last Name', 'First Name',
        'Second Name', 'Grade', 'Grade Percent']
    skip_tail = [
        'Cohort Name', 'Enrollment Track', 'Verification Status',
        'Certificate Eligible', 'Certificate Delivered', 'Certificate Type']

    cols = [(i, worksheet.cell(column=(i+1), row=1).value)
           for i in range(worksheet.max_column)]

    while cols and cols[0][1] in skip_head:
        cols.pop(0)

    while cols and cols[-1][1] in skip_tail:
        cols.pop()

    return [i for (i, s) in cols if 'avg' not in s.lower()]


def extract_name_edx_or_urfu(row, legend):
    name = legend.get_item(row, 'full_name', dtype=str)
    if not name:
        name = ' '.join([
            legend.get_item(row, 'First Name', dtype=str),
            legend.get_item(row, 'Last Name', dtype=str)])

    parts = name.split(maxsplit=1)
    return (parts + ['', ''])[:2]


def process_edx_or_urfu(worksheet, students, export):
    tasks = get_tasks_edx_urfu(worksheet)
    legend = None

    platform_id = str(uuid.uuid4())

    for (i, row) in enumerate(worksheet.rows):
        if i == 0:
            legend = WorksheetLegend(row)
            continue

        (student_id, new) = students.get_or_create(
            legend.get_item(row, 'id', 'Student ID'),
            legend.get_item(row, 'username', 'Username')
        )

        if new:
            s0 = export.get_sheet_by_name('Опросник студента подход 2')
            name, surname = extract_name_edx_or_urfu(row, legend)
            s0.append([
                student_id, "", "", "", "", "", "", "",
                legend.get_item(row, 'email', 'Email'),
                surname, name,
                legend.get_item(row, 'second_name', 'Second Name')])

            s1 = export.get_sheet_by_name('Участник обучения')
            s1.append([student_id])

        student_course_id = str(uuid.uuid4())
        s2 = export.get_sheet_by_name('Участн-Курс-Дисциплина')
        s2.append([
            student_course_id, student_id, "", platform_id, "",
            legend.get_item(row, 'username', 'Username'),
            str(legend.get_item(row, 'id', 'Student ID', dtype=int))])

        result_id = str(uuid.uuid4())
        s3 = export.get_sheet_by_name('Результаты обучения на курсе')
        s3.append([
            result_id, student_course_id, "", "",
            legend.get_item(row, 'grade', 'Grade')])

        s4 = export.get_sheet_by_name('Текущая успеваемость')
        for task in tasks:
            current_score_id = str(uuid.uuid4())
            s4.append([
                current_score_id, result_id, legend.names[task], "", "", "",
                check_value(row[task].value, float)])

    return platform_id


def get_weight(name):
    return convert_value(name.rsplit('/', 1)[-1].replace(',', '.'), float) or 1


def get_tasks_moodle(worksheet):
    skip_head = [
        'Фамилия', 'Имя', 'Учреждение (организация)', 'Отдел',
        'Адрес электронной почты', 'Состояние', 'Тест начат', 'Завершено',
        'Затраченное время', 'Индивидуальный номер', 'First name', 'Surname',
        'ID number', 'Institution', 'Department', 'Email address', 'State',
        'Started on', 'Completed', 'Time taken']
    skip_tail = [
        'Последние загруженные из этого курса',
        'Last downloaded from this course']
    skip_middle = [
        'Оценка', 'Grade', 'Course total', 'Итоговая оценка за курс']

    cols = [(i, worksheet.cell(column=(i+1), row=1).value)
            for i in range(worksheet.max_column)]

    while cols and cols[0][1] in skip_head:
        cols.pop(0)

    while cols and cols[-1][1] in skip_tail:
        cols.pop()

    return [(i, get_weight(s)) for (i, s) in cols
            if all(x not in s for x in skip_middle)]


def get_total_column_moodle(heads):
    names = ['Оценка', 'Grade', 'Course total', 'Итоговая оценка за курс']

    idxs = [i for (i, s) in enumerate(heads) if any(x in s for x in names)]
    if not idxs:
        return None

    return (idxs[0], get_weight(heads[idxs[0]]))



def process_moodle(worksheet, students, export):
    def normalize(value, weight):
        try:
            return float(value/weight)
        except Exception:
            return ""

    def fix_name(name):
        return re.sub(r'\(.*\)', '', name.rsplit('/', 1)[0]).strip()

    tasks = get_tasks_moodle(worksheet)
    legend = None
    total = None

    platform_id = str(uuid.uuid4())

    for (i, row) in enumerate(worksheet.rows):
        if i == 0:
            legend = WorksheetLegend(row)
            total = get_total_column_moodle(legend.names)
            continue

        if row[0].value in ('Overall average', 'Общее среднее'):
            continue

        (student_id, new) = students.get_or_create(
            legend.get_item(row, 'Имя', 'First name'),
            legend.get_item(row, 'Фамилия', 'Surname')
        )
        if new:
            s0 = export.get_sheet_by_name('Опросник студента подход 2')
            name = legend.get_item(row, 'Имя', 'First name')
            surname = legend.get_item(row, 'Фамилия', 'Surname')
            s0.append([
                student_id, "", "", "", "", "", "", "",
                legend.get_item(
                    row, 'Адрес электронной почты', 'Email address'),
                surname, name])

            s1 = export.get_sheet_by_name('Участник обучения')
            s1.append([student_id])

        student_course_id = str(uuid.uuid4())
        s2 = export.get_sheet_by_name('Участн-Курс-Дисциплина')
        s2.append([
            student_course_id, student_id, "", platform_id, "", "",
            legend.get_item(row, 'Индивидуальный номер', 'ID number')])

        result_id = str(uuid.uuid4())
        s3 = export.get_sheet_by_name('Результаты обучения на курсе')
        s3.append([
            result_id, student_course_id, "", "",
            normalize(row[total[0]].value, total[1])])

        s4 = export.get_sheet_by_name('Текущая успеваемость')
        for (task, weight) in tasks:
            current_score_id = str(uuid.uuid4())
            s4.append([
                current_score_id, result_id, fix_name(legend.names[task]),
                "", legend.get_item(row, 'Тест начат', 'Завершено'),
                legend.get_item(row, 'Started on', 'Completed'),
                normalize(row[task].value, weight)])

    return platform_id


def main(args):
    workbooks = load_workbooks(['edx', 'moodle', 'urfu'])

    export = create_export_skeleton()
    students = StudentsFactory()

    for ((folder, name), workbook) in workbooks:
        print('Processing "{}"...'.format(name))
        s0 = export.get_sheet_by_name('Дисциплина с платформы')
        if folder in ('edx', 'urfu'):
            platform_id = process_edx_or_urfu(
                workbook.worksheets[0], students, export)

            s0.append([platform_id, name.split('_')[1]])
        elif folder in ('moodle',):
            platform_id = process_moodle(
                workbook.worksheets[0], students, export)
            s0.append([platform_id])

    export.save(filename='export.xlsx')


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))

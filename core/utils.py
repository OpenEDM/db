import random
from collections import OrderedDict
from io import BytesIO, StringIO

from django.db.models import NullBooleanField, ForeignKey, CharField, TextField, IntegerField, PositiveIntegerField, \
    DateTimeField, DateField, BooleanField, DecimalField
from django.db.transaction import atomic
from django.utils.crypto import get_random_string
from django.utils.timezone import now
from openpyxl import Workbook, load_workbook


class DataModel:
    def __init__(self) -> None:
        super().__init__()
        self._load_data_model()

    def get_models(self):
        from core.models import Student, StudentData1, Teacher, SemestrData, Activity, OOP, SubjectUP, \
            SubjectPlatform, StudentCourseSubject, Result, CurrentProgress, Quiz, StudentData2

        return Student, StudentData1, Teacher, SemestrData, Activity, OOP, SubjectUP, SubjectPlatform, \
            StudentCourseSubject, Result, CurrentProgress, Quiz, StudentData2

    def _load_data_model(self):
        self.data_model = OrderedDict()
        self.reverse_data_model = OrderedDict()

        for model in self.get_models():
            fields = OrderedDict(model.Mapping.fields)
            self.data_model[model.Mapping.name] = {
                'model': model,
                'fields': fields
            }

            self.reverse_data_model[model] = {
                'name': model.Mapping.name,
                'fields': OrderedDict((y, x) for x, y in fields.items())
            }


def _get_field(model, field_name):
    for _field in model._meta.get_fields():
        if _field.name == field_name:
            return _field
    return None


class Loader:
    @classmethod
    def add_prefix(cls, prefix, value):
        return '%s_%s' % (prefix, value)

    def get_value(self, model, field_name, prefix, raw_value):
        if field_name == 'import_id':
            raw_value = self.add_prefix(prefix, raw_value)
        field = _get_field(model, field_name)
        print(field_name, type(field))
        if isinstance(field, ForeignKey):
            related_model = field.related_model
            return related_model.objects.filter(import_id=self.add_prefix(prefix, raw_value)).first()
        elif isinstance(field, NullBooleanField):
            return raw_value
        return raw_value

    @atomic
    def load(self, prefix, filename):
        data_model = DataModel().data_model
        wb = load_workbook(filename, read_only=True)
        for sheet_name, mapping_data in data_model.items():
            model = mapping_data['model']
            ws = wb.get_sheet_by_name(sheet_name)
            header = [cell.value for cell in ws[1]]
            fields = [mapping_data['fields'].get(col_name, None) for col_name in header]
            for row_num, row in enumerate(ws.rows):
                if row_num < 1:
                    continue
                data = {
                    fields[cell_i]: self.get_value(model, fields[cell_i], prefix, cell.value)
                    for cell_i, cell in enumerate(row)
                }
                data.pop(None, None)
                obj = model.objects.update_or_create(import_id=data.pop('import_id'), defaults=data)
                print(obj)


class Exporter:
    @classmethod
    def remove_prefix(cls, value):
        return value.split('_', maxsplit=1)[-1]

    def get_value(self, model, field_name, obj):
        field = _get_field(model, field_name)
        value = getattr(obj, field_name)
        if field_name == 'import_id':
            return self.remove_prefix(value)
        if isinstance(field, ForeignKey):
            value = self.remove_prefix(value.import_id)
        return value

    def export(self):
        reverse_data_model = DataModel().reverse_data_model
        wb = Workbook(write_only=True)
        for model, mapping_data in reverse_data_model.items():
            ws = wb.create_sheet(title=mapping_data['name'])
            ws.append(list(mapping_data['fields'].values()))
            for obj in model.objects.all():
                ws.append([self.get_value(model, field, obj) for field in list(mapping_data['fields'].keys())])
        return wb


class DataGenerator:
    def generate_value(self, model, field_name):
        field = _get_field(model, field_name)
        if isinstance(field, (CharField, TextField)):
            return get_random_string(3)
        elif isinstance(field, (IntegerField, PositiveIntegerField)):
            return random.randint(0, 1000)
        elif isinstance(field, DecimalField):
            return random.randint(0, 1000) * 0.3
        elif isinstance(field, ForeignKey):
            return field.related_model.objects.order_by('?').first()
        elif isinstance(field, DateTimeField):
            return now()
        elif isinstance(field, DateField):
            return now().date()
        elif isinstance(field, (NullBooleanField, BooleanField)):
            return random.choice([True, False])
        print(model, field_name)

    @atomic
    def generate(self):
        reverse_data_model = DataModel().reverse_data_model
        for model, mapping_data in reverse_data_model.items():
            model.objects.all().delete()

            start_index = random.randint(1, 10000)
            for i in range(start_index, start_index + 100):
                data = {field: self.generate_value(model, field) for field in list(mapping_data['fields'].keys())}
                data['import_id'] = Loader.add_prefix('urfu', i)
                model.objects.create(**data)
